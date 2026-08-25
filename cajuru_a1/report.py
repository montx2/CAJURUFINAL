"""Relatório legível (HTML) e planilha de diagnóstico do processamento."""

from __future__ import annotations

import html
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from cajuru_a1.cnpjutil import format_cnpj

_STATUS_COLORS = {
    "PRONTO": "2e7d32",
    "SUBSTITUIDO": "f9a825",
    "SEM_SENHA": "c62828",
    "CONFLITO": "c62828",
    "VENCIDO": "ef6c00",
    "DUPLICADO": "6a1b9a",
    "SEM_CNPJ": "c62828",
    "AINDA_NAO_VALIDO": "ef6c00",
}


def write_html_report(result, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "<!DOCTYPE html><html lang='pt-BR'><head><meta charset='utf-8'>",
        "<title>Relatório Cajuru A1</title>",
        "<style>",
        "body{font-family:Segoe UI,Arial,sans-serif;margin:24px;color:#222}",
        "h1{font-size:20px} h2{font-size:16px;margin-top:24px}",
        "table{border-collapse:collapse;width:100%;font-size:13px;margin-top:8px}",
        "th,td{border:1px solid #ccc;padding:6px 8px;text-align:left;vertical-align:top}",
        "th{background:#f0f0f0}",
        ".kpi{display:inline-block;margin-right:18px;padding:10px 16px;background:#f5f5f5;border-radius:6px}",
        ".kpi b{font-size:20px;display:block}",
        ".badge{padding:2px 8px;border-radius:10px;color:#fff;font-size:11px;font-weight:bold}",
        "</style></head><body>",
        f"<h1>Relatório Cajuru A1 — {datetime.now().strftime('%d/%m/%Y %H:%M')}</h1>",
        "<div>",
    ]
    stats = result.stats
    kpis = [
        ("Total de arquivos", stats.get("total_arquivos", 0)),
        ("Abertos com senha", stats.get("abertos", 0)),
        ("Prontos (selecionados)", stats.get("selecionados", 0)),
        ("Duplicatas/substituídos", stats.get("duplicatas", 0)),
        ("Rejeitados", stats.get("rejeitados", 0)),
    ]
    for label, value in kpis:
        lines.append(f"<span class='kpi'><b>{value}</b>{html.escape(str(label))}</span>")
    lines.append("</div>")

    # Selecionados
    lines.append("<h2>Certificados prontos para exportação</h2>")
    lines.append("<table><tr><th>CNPJ</th><th>Empresa</th><th>Arquivo</th><th>Validade</th><th>Origem da senha</th></tr>")
    for doc, cert in sorted(result.selected.items()):
        validade = cert.not_after.strftime("%d/%m/%Y") if cert.not_after else "—"
        lines.append(
            f"<tr><td>{html.escape(format_cnpj(doc))}</td>"
            f"<td>{html.escape(cert.company_from_cert or '')}</td>"
            f"<td>{html.escape(cert.filename)}</td>"
            f"<td>{validade}</td>"
            f"<td>{html.escape(cert.password_source or '')}</td></tr>"
        )
    lines.append("</table>")

    # Rejeitados / duplicados
    rows = []
    for cert, code, reason in result.rejected:
        rows.append((cert, code, reason))
    for cert in result.duplicates:
        rows.append((cert, "SUBSTITUIDO", cert.extra.get("motivo_substituicao", "Substituído por cópia mais nova")))
    if rows:
        lines.append("<h2>Não exportados</h2>")
        lines.append("<table><tr><th>Status</th><th>Arquivo</th><th>CNPJ no nome</th><th>Empresa</th><th>Motivo</th></tr>")
        for cert, code, reason in rows:
            color = _STATUS_COLORS.get(code, "555")
            if cert is None:
                lines.append(f"<tr><td><span class='badge' style='background:{color}'>{code}</span></td><td colspan='4'>{html.escape(reason)}</td></tr>")
                continue
            lines.append(
                f"<tr><td><span class='badge' style='background:{color}'>{code}</span></td>"
                f"<td>{html.escape(cert.filename)}</td>"
                f"<td>{html.escape(format_cnpj(cert.cnpj_filename) if cert.cnpj_filename else '')}</td>"
                f"<td>{html.escape(cert.company_from_cert or '')}</td>"
                f"<td>{html.escape(reason)}</td></tr>"
            )
        lines.append("</table>")

    # Avisos do vault
    if result.vault and result.vault.findings:
        lines.append("<h2>Avisos das planilhas de senha</h2><ul>")
        for finding in result.vault.findings:
            lines.append(f"<li><b>{html.escape(finding.get('code',''))}</b>: {html.escape(finding.get('message',''))}</li>")
        lines.append("</ul>")

    lines.append("</body></html>")
    output_path.write_text("\n".join(lines), encoding="utf-8")
    return output_path


def write_xlsx_report(result, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Prontos"
    headers = ["CNPJ", "Empresa", "Arquivo original", "Validade", "Origem da senha", "SHA-256"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for doc, cert in sorted(result.selected.items()):
        ws.append([
            format_cnpj(doc),
            cert.company_from_cert or "",
            cert.filename,
            cert.not_after.strftime("%d/%m/%Y") if cert.not_after else "",
            cert.password_source or "",
            cert.sha256,
        ])

    ws2 = wb.create_sheet("Rejeitados")
    ws2.append(["Status", "Arquivo", "CNPJ no nome", "Empresa", "Motivo", "Validade"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    rows = list(result.rejected) + [(c, "SUBSTITUIDO", c.extra.get("motivo_substituicao", "")) for c in result.duplicates]
    for cert, code, reason in rows:
        if cert is None:
            ws2.append([code, "", "", "", reason, ""])
            continue
        ws2.append([
            code,
            cert.filename,
            format_cnpj(cert.cnpj_filename) if cert.cnpj_filename else "",
            cert.company_from_cert or "",
            reason,
            cert.not_after.strftime("%d/%m/%Y") if cert.not_after else "",
        ])

    ws3 = wb.create_sheet("Resumo")
    ws3.append(["Métrica", "Valor"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for key, value in result.stats.items():
        ws3.append([key, value])

    wb.save(output_path)
    return output_path
