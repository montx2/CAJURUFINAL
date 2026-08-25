"""Geração do pacote de importação Jettax 360.

Regras que corrigem os erros do importador:
- cada PFX dentro do ZIP chama-se EXATAMENTE ``<CNPJ>.pfx`` (14 dígitos);
- a planilha modelo oficial tem UMA linha por CNPJ, sem duplicatas;
- CNPJ e senha são preenchidos como texto para não perder zero à esquerda;
- arquivos que não puderam ser abertos/identificados vão para um relatório de
  rejeitados com o motivo, NÃO entram no ZIP nem na planilha.
"""

from __future__ import annotations

import csv
import os
import shutil
import stat
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl

from cajuru_a1.cnpjutil import format_cnpj, is_valid_doc, only_digits, pad_cnpj

TEMPLATE_PATH = Path(__file__).parent / "resources" / "modelo_import_certificados.xlsx"
TEMPLATE_SHEET = "Certificados"
COL_CNPJ = 1
COL_SENHA = 2


def _clear_datalines(sheet) -> None:
    """Apaga linhas de dados preservando cabeçalho e formatação do modelo."""
    for row in range(2, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, col)
            cell.value = None
            try:
                cell.hyperlink = None
            except (AttributeError, TypeError):
                pass


def build_bundle(
    selected: dict,
    output_dir: str | Path,
    *,
    template_path: str | Path | None = None,
    senha_manual: bool = True,
    rejeitados: list | None = None,
) -> dict:
    """Cria ``<output>/lotes/lote_<timestamp>/`` com ZIP, planilha e relatórios.

    ``selected`` é um dicionário ``{cnpj: PfxInfo}`` (saída de ``run_pipeline``).
    """
    template_path = Path(template_path) if template_path else TEMPLATE_PATH
    if not template_path.is_file():
        raise FileNotFoundError(
            f"Modelo oficial de importação não encontrado em {template_path}."
        )

    output_dir = Path(output_dir).expanduser().resolve(strict=False)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    bundle_dir = output_dir / "lotes" / f"lote_{stamp}"
    bundle_dir.mkdir(parents=True, exist_ok=False)
    try:
        bundle_dir.chmod(stat.S_IRWXU)
    except OSError:
        pass

    # Valida estrita: um CNPJ válido por certificado, sem repetição.
    elegiveis: list[tuple[str, object]] = []
    vistos: set[str] = set()
    for doc, cert in sorted(selected.items()):
        document = pad_cnpj(only_digits(doc))
        if not is_valid_doc(document):
            continue
        if document in vistos:
            raise RuntimeError(f"CNPJ duplicado no lote: {format_cnpj(document)}")
        source = Path(cert.source_path)
        if not source.is_file():
            raise FileNotFoundError(f"Arquivo de certificado não encontrado: {source}")
        vistos.add(document)
        elegiveis.append((document, cert))

    if not elegiveis:
        shutil.rmtree(bundle_dir, ignore_errors=True)
        raise RuntimeError("Nenhum certificado válido e elegível para exportar.")

    zip_path = bundle_dir / "certificados_jettax.zip"
    planilha_path = bundle_dir / "planilha_importacao_jettax.xlsx"
    csv_senhas = bundle_dir / "senhas_para_conferencia.csv"
    rejeitados_path = bundle_dir / "rejeitados.csv"

    try:
        # 1) ZIP — cada arquivo nomeado EXATAMENTE como o CNPJ (14 dígitos).pfx
        _build_zip(elegiveis, zip_path)

        # 2) Planilha modelo OFICIAL do Jettax preenchida.
        _build_planilha(elegiveis, planilha_path, template_path, senha_manual=senha_manual)

        # 3) CSV de conferência de senhas (NUNCA é a planilha enviada; é apoio).
        _build_senhas_csv(elegiveis, csv_senhas)

        # 4) Relatório de rejeitados com motivo.
        if rejeitados:
            _build_rejeitados_csv(rejeitados, rejeitados_path)

        _write_readme(bundle_dir, len(elegiveis), zip_path, planilha_path,
                      csv_senhas, rejeitados_path if rejeitados else None,
                      senha_manual=senha_manual)

        for path in (zip_path, planilha_path, csv_senhas):
            try:
                path.chmod(stat.S_IRUSR | stat.S_IWUSR)
            except OSError:
                pass
    except Exception:
        shutil.rmtree(bundle_dir, ignore_errors=True)
        raise

    return {
        "dir": bundle_dir,
        "zip": zip_path,
        "planilha": planilha_path,
        "csv_senhas": csv_senhas,
        "rejeitados": rejeitados_path if rejeitados else None,
        "quantidade": len(elegiveis),
    }


def _build_zip(elegiveis, zip_path: Path) -> None:
    """ZIP com arquivos nomeados <CNPJ>.pfx — requisito do importador Jettax."""
    used: set[str] = set()
    with zipfile.ZipFile(zip_path, "x", zipfile.ZIP_DEFLATED) as archive:
        for document, cert in elegiveis:
            name = f"{document}.pfx"  # ex.: 12345678000195.pfx
            if name in used:
                raise RuntimeError(f"Nome duplicado no ZIP: {name}")
            used.add(name)
            archive.write(cert.source_path, name)


def _build_planilha(elegiveis, planilha_path: Path, template_path: Path, *,
                    senha_manual: bool) -> None:
    workbook = openpyxl.load_workbook(template_path)
    if TEMPLATE_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"O modelo não tem a aba obrigatória '{TEMPLATE_SHEET}'")
    sheet = workbook[TEMPLATE_SHEET]
    _clear_datalines(sheet)

    for index, (document, cert) in enumerate(elegiveis, start=2):
        cnpj_cell = sheet.cell(index, COL_CNPJ)
        cnpj_cell.value = format_cnpj(document)
        cnpj_cell.data_type = "s"
        senha_cell = sheet.cell(index, COL_SENHA)
        senha_cell.value = "" if senha_manual or cert.password is None else str(cert.password)
        senha_cell.data_type = "s"

    temporary = planilha_path.with_suffix(".tmp.xlsx")
    workbook.save(temporary)
    workbook.close()
    os.replace(temporary, planilha_path)


def _build_senhas_csv(elegiveis, csv_path: Path) -> None:
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow(["CNPJ", "EMPRESA", "ARQUIVO_ORIGINAL", "SENHA_CERTIFICADO", "VALIDADE"])
        for document, cert in elegiveis:
            validade = cert.not_after.strftime("%d/%m/%Y") if getattr(cert, "not_after", None) else ""
            writer.writerow([
                format_cnpj(document),
                cert.company_from_cert or "",
                cert.filename,
                "" if cert.password is None else str(cert.password),
                validade,
            ])


def _build_rejeitados_csv(rejeitados, csv_path: Path) -> None:
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow(["ARQUIVO", "CODIGO", "MOTIVO", "CNPJ_NO_NOME", "EMPRESA"])
        for cert, code, reason in rejeitados:
            if cert is None:
                writer.writerow(["", code, reason, "", ""])
            else:
                writer.writerow([
                    getattr(cert, "filename", ""),
                    code,
                    reason,
                    format_cnpj(cert.cnpj_filename) if getattr(cert, "cnpj_filename", None) else "",
                    getattr(cert, "company_from_cert", "") or "",
                ])


def _write_readme(bundle_dir: Path, quantidade: int, zip_path: Path,
                  planilha_path: Path, csv_senhas: Path,
                  rejeitados_path: Path | None, *, senha_manual: bool) -> None:
    lines = [
        "LOTE DE IMPORTACAO JETTAX 360",
        "==============================",
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        f"Certificados no lote: {quantidade}",
        "",
        "ARQUIVOS PARA IMPORTAR NO JETTAX (Clientes > Importar):",
        f"  1. {zip_path.name}",
        f"     - cada certificado esta nomeado como <CNPJ>.pfx (14 digitos)",
        f"  2. {planilha_path.name}",
        f"     - modelo oficial; coluna SENHA {'em branco (preencha voce)' if senha_manual else 'preenchida'}",
        "",
        "APOIO (nao envie ao Jettax):",
        f"  - {csv_senhas.name}: senhas validadas para conferencia",
    ]
    if rejeitados_path:
        lines.append(f"  - {rejeitados_path.name}: arquivos que NAO entraram no lote e o motivo")
    lines += [
        "",
        "COMO IMPORTAR:",
        "  1. No Jettax, abra Clientes > Importar.",
        "  2. Selecione o ZIP e a planilha juntos.",
        "  3. Confira cada CNPJ e digite a senha do A1 (veja o CSV de apoio).",
        "  4. Conclua a importacao.",
        "",
        "Estes arquivos contem senhas. Guarde em local seguro e apague depois.",
    ]
    (bundle_dir / "LEIA-ME.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
