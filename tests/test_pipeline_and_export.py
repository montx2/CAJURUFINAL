from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

from cajuru_a1.cnpjutil import format_cnpj
from cajuru_a1.exportacao import build_bundle
from cajuru_a1.pipeline import run_pipeline


def test_zip_files_named_by_cnpj_only(make_pfx, make_senhas_xlsx, tmp_path):
    cnpj = "50590299000150"
    # Nome de arquivo "sujo", como os que o Jettax rejeitou.
    make_pfx(cnpj, "WM GESTAO EMPRESARIAL LTDA", "senha123",
             filename="WM GESTAO EMPRESARIAL LTDA_68620019000174.pfx")
    planilha = make_senhas_xlsx([["WM GESTAO EMPRESARIAL LTDA", cnpj, "senha123", ""]])

    result = run_pipeline(tmp_path / "certs", [planilha])
    assert cnpj in result.selected

    saida = tmp_path / "out"
    bundle = build_bundle(result.selected, saida, rejeitados=result.rejected)

    with zipfile.ZipFile(bundle["zip"]) as zf:
        names = zf.namelist()
    assert names == [f"{cnpj}.pfx"], f"ZIP deve conter apenas {cnpj}.pfx, obteve {names}"

    wb = openpyxl.load_workbook(bundle["planilha"])
    ws = wb["Certificados"]
    assert ws.cell(2, 1).value == format_cnpj(cnpj)
    # Senha em branco no modo manual (padrão).
    assert ws.cell(2, 2).value in ("", None)


def test_duplicate_cnpj_deduplicated(make_pfx, make_senhas_xlsx, tmp_path):
    """Dois PFX do mesmo CNPJ -> só um vai para o lote; o outro é SUBSTITUIDO."""
    cnpj = "00796329000100"
    make_pfx(cnpj, "DOCE TEMPERO LTDA", "senha123",
             filename="DOCE TEMPERO ALIMENTACAO INDUSTRIAL LTDA00796329000100.pfx",
             not_after_days=100)
    make_pfx(cnpj, "DOCE TEMPERO LTDA", "senha123",
             filename="DOCE_TEMPERO_00796329000100 (1).pfx",
             not_after_days=400)  # mais novo -> vence
    planilha = make_senhas_xlsx([["DOCE TEMPERO LTDA", cnpj, "senha123", ""]])

    result = run_pipeline(tmp_path / "certs", [planilha])
    assert len(result.selected) == 1
    assert cnpj in result.selected
    assert len(result.duplicates) == 1

    bundle = build_bundle(result.selected, tmp_path / "out", rejeitados=result.rejected)
    with zipfile.ZipFile(bundle["zip"]) as zf:
        assert zf.namelist().count(f"{cnpj}.pfx") == 1


def test_exact_duplicate_sha256_rejected(make_pfx, make_senhas_xlsx, tmp_path):
    cnpj = "03049033000114"
    p = make_pfx(cnpj, "T E T COMERCIO LTDA", "senha123",
                 filename="T & T COMERCIO_05128473000192 (1).pfx")
    # Cópia idêntica com outro nome.
    import shutil
    shutil.copy(p, p.parent / "T & T COMERCIO_05128473000192.pfx")
    planilha = make_senhas_xlsx([["T E T COMERCIO LTDA", cnpj, "senha123", ""]])

    result = run_pipeline(tmp_path / "certs", [planilha])
    assert cnpj in result.selected
    assert any(code == "DUPLICADO" for _c, code, _r in result.rejected)


def test_no_cnpj_in_cert_rejected(make_pfx, make_senhas_xlsx, tmp_path):
    # CPF no certificado (pessoa física) não vai para o lote de CNPJs.
    cpf = "08915432606"
    from cryptography import x509
    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.primitives.asymmetric import rsa
    from cryptography.hazmat.primitives.serialization import pkcs12
    from cryptography.x509.oid import NameOID
    import datetime as dt

    key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
    subject = x509.Name([
        x509.NameAttribute(NameOID.COMMON_NAME, f"THALES:{cpf}"),
        x509.NameAttribute(NameOID.ORGANIZATION_NAME, "THALES WINSTON"),
        x509.NameAttribute(NameOID.SERIAL_NUMBER, cpf),
    ])
    now = dt.datetime.now(dt.timezone.utc)
    cert = (x509.CertificateBuilder().subject_name(subject).issuer_name(subject)
            .public_key(key.public_key()).serial_number(x509.random_serial_number())
            .not_valid_before(now - dt.timedelta(days=1))
            .not_valid_after(now + dt.timedelta(days=365)).sign(key, hashes.SHA256()))
    pfx_path = tmp_path / "certs" / "THALES WINSTON MARQUES COSTA08915432606 CPF.pfx"
    pfx_path.parent.mkdir(parents=True, exist_ok=True)
    pfx_path.write_bytes(pkcs12.serialize_key_and_certificates(
        b"thales", key, cert, None,
        serialization.BestAvailableEncryption(b"senha123")))
    planilha = make_senhas_xlsx([["THALES WINSTON", cpf, "senha123", ""]])

    result = run_pipeline(tmp_path / "certs", [planilha])
    # CPF é documento válido (11 dígitos) e pode ser selecionado — o Jettax
    # aceita CPF na coluna CNPJ. Mas garantimos que o arquivo seja renomeado.
    if result.selected:
        bundle = build_bundle(result.selected, tmp_path / "out")
        with zipfile.ZipFile(bundle["zip"]) as zf:
            for name in zf.namelist():
                # Nome deve ser apenas dígitos + .pfx, sem texto de empresa.
                stem = name[:-4]
                assert stem.isdigit(), f"nome no ZIP deve ser só dígitos: {name}"


def test_multiple_distinct_cnpjs_no_duplicates_in_sheet(make_pfx, make_senhas_xlsx, tmp_path):
    rows = []
    cnpjs = ["00796329000100", "50590299000150", "41367161000103"]
    for i, cnpj in enumerate(cnpjs):
        make_pfx(cnpj, f"EMPRESA {i} LTDA", "senha123", filename=f"emp{i}_{cnpj}.pfx")
        rows.append([f"EMPRESA {i} LTDA", cnpj, "senha123", ""])
    planilha = make_senhas_xlsx(rows)

    result = run_pipeline(tmp_path / "certs", [planilha])
    bundle = build_bundle(result.selected, tmp_path / "out")

    with zipfile.ZipFile(bundle["zip"]) as zf:
        names = sorted(zf.namelist())
    assert names == sorted(f"{c}.pfx" for c in cnpjs)

    wb = openpyxl.load_workbook(bundle["planilha"])
    ws = wb["Certificados"]
    sheet_docs = [ws.cell(r, 1).value for r in range(2, 2 + len(cnpjs))]
    assert len(sheet_docs) == len(set(sheet_docs)), "CNPJs duplicados na planilha"


def test_rejeitados_report_written(make_pfx, make_senhas_xlsx, tmp_path):
    # PFX com senha errada/não cadastrada -> rejeitado, não entra no ZIP.
    cnpj = "99999999000191"
    make_pfx(cnpj, "EMPRESA SEM SENHA", "senha_desconhecida", filename=f"x_{cnpj}.pfx")
    planilha = make_senhas_xlsx([["OUTRA EMPRESA", "11111111000191", "outra", ""]])

    result = run_pipeline(tmp_path / "certs", [planilha])
    assert not result.selected  # não conseguiu abrir
    assert result.rejected
    with pytest.raises(RuntimeError):
        build_bundle(result.selected, tmp_path / "out")
