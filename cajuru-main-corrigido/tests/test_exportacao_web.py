"""Exportação local de certificados A1 sem acesso ao Jettax."""

from __future__ import annotations

import csv
import zipfile

import openpyxl

from cajuru_a1.exportacao import export_all_opened
from cajuru_a1.passwords import PasswordVault, candidate_passwords
from cajuru_a1.pfx import PfxInfo, inspect_file

from conftest import make_pfx


def _opened(path, *, password: str, company: str):
    candidates = candidate_passwords(
        vault=PasswordVault(), empresa=company, cnpj="12345678000195", include_empty=True
    ) + [(password, "teste")]
    cert = inspect_file(path, path, candidates)
    assert cert.opened
    return cert


def test_exporta_zip_jettax_com_cnpj_puro_e_planilha_correspondente(tmp_path):
    source = tmp_path / "origem"
    source.mkdir()
    # Simula exatamente o nome que o Jettax rejeitava: razão social + CNPJ e
    # extensão .p12. O nome dentro do ZIP precisa ser somente CNPJ.pfx.
    pfx = source / "CENTRO DE RADIOLOGIA LTDA_12345678000195.p12"
    make_pfx(pfx, password="senha-validada", cnpj="12345678000195", company="EMPRESA TESTE")
    cert = _opened(pfx, password="senha-validada", company="EMPRESA TESTE")

    export = export_all_opened([cert], tmp_path / "saida")
    assert export["quantidade"] == 1
    assert export["quantidade_revisao"] == 0
    assert export["revisao"] is None
    assert export["zip"] and export["zip"].is_file()
    with zipfile.ZipFile(export["zip"]) as archive:
        assert archive.namelist() == ["12345678000195.pfx"]

    assert export["senhas"] and export["senhas"].is_file()
    with export["senhas"].open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=";"))
    assert rows[1][0] == "12345678000195.pfx"
    assert rows[1][3] == "senha-validada"

    # A exportação também gera a planilha modelo OFICIAL do Jettax (aba
    # "Certificados") com o mesmo CNPJ do ZIP e a senha preenchida.
    assert export.get("planilha") and export["planilha"].is_file()
    workbook = openpyxl.load_workbook(export["planilha"])
    assert workbook.sheetnames == ["Leia-me", "Certificados", "Regimes"]
    sheet = workbook["Certificados"]
    assert sheet.cell(1, 1).value == "CNPJ *"
    assert sheet.cell(1, 2).value == "Senha Certificado *"
    assert sheet.cell(2, 1).value == "12.345.678/0001-95"
    assert sheet.cell(2, 2).value == "senha-validada"
    # Demais colunas do modelo permanecem em branco (sem o exemplo residual).
    assert all(sheet.cell(2, col).value in (None, "") for col in range(3, sheet.max_column + 1))
    workbook.close()


def test_exportacao_separa_cpf_e_duplicata_do_zip_pronto_para_jettax(tmp_path):
    source = tmp_path / "origem"
    source.mkdir()
    old_path = source / "EMPRESA ANTIGA_12345678000195.pfx"
    new_path = source / "EMPRESA NOVA_12345678000195.pfx"
    make_pfx(old_path, password="senha-antiga", cnpj="12345678000195", company="EMPRESA ANTIGA", days_valid=30)
    make_pfx(new_path, password="senha-nova", cnpj="12345678000195", company="EMPRESA NOVA", days_valid=400)
    old = _opened(old_path, password="senha-antiga", company="EMPRESA ANTIGA")
    new = _opened(new_path, password="senha-nova", company="EMPRESA NOVA")

    # Simula um A1 de pessoa física que abriu normalmente, mas não pode entrar
    # em um importador que exige CNPJ no nome do arquivo.
    cpf_path = source / "PAULA_FONSECA_07115478660.pfx"
    cpf_path.write_bytes(new_path.read_bytes())
    cpf = PfxInfo(
        source_path=str(cpf_path),
        temp_path=str(cpf_path),
        filename=cpf_path.name,
        sha256="cpf",
        size=cpf_path.stat().st_size,
        cnpj_cert="07115478660",
        company_from_cert="PAULA FONSECA",
        password="senha-cpf",
        opened=True,
    )

    export = export_all_opened([old, new, cpf], tmp_path / "saida")
    assert export["quantidade"] == 1
    assert export["quantidade_revisao"] == 2
    with zipfile.ZipFile(export["zip"]) as archive:
        assert archive.namelist() == ["12345678000195.pfx"]
    with export["senhas"].open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=";"))
    # O certificado mais novo vence mais tarde e é o que fica no ZIP principal.
    assert rows[1][2] == "EMPRESA NOVA"
    assert rows[1][3] == "senha-nova"

    assert export["revisao"] and export["revisao"].is_file()
    with zipfile.ZipFile(export["revisao"]) as archive:
        assert sorted(archive.namelist()) == sorted([old_path.name, cpf_path.name])

    with export["nao_exportados"].open(encoding="utf-8-sig", newline="") as handle:
        skipped = list(csv.reader(handle, delimiter=";"))
    reasons = "\n".join(row[1] for row in skipped[1:])
    assert "CPF" in reasons
    assert "CNPJ duplicado" in reasons


def test_exportacao_nao_cria_zip_jettax_vazio_quando_so_ha_cpf(tmp_path):
    source = tmp_path / "origem"
    source.mkdir()
    cpf_path = source / "PESSOA_FISICA_07115478660.pfx"
    cpf_path.write_bytes(b"certificado-de-revisao")
    cpf = PfxInfo(
        source_path=str(cpf_path),
        temp_path=str(cpf_path),
        filename=cpf_path.name,
        sha256="cpf-only",
        size=cpf_path.stat().st_size,
        cnpj_cert="07115478660",
        password="senha-cpf",
        opened=True,
    )

    export = export_all_opened([cpf], tmp_path / "saida")
    assert export["quantidade"] == 0
    assert export["zip"] is None
    assert export["senhas"] is None
    assert export["planilha"] is None
    assert export["revisao"] and export["revisao"].is_file()
    assert not (export["dir"] / "todos_certificados_a1.zip").exists()
    assert "CNPJ interno válido" in (export["dir"] / "planilha_nao_gerada.txt").read_text(encoding="utf-8")
