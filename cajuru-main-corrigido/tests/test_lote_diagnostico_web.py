"""Testes do lote manual, diagnóstico e conciliação segura."""

from __future__ import annotations

import pytest

from cajuru_a1.diagnostico import (
    build_diagnostico, write_diagnostico_excel, write_diagnostico_html,
)
from cajuru_a1.lote import (
    build_importacao_jettax, build_persistent_bundle, build_senhas_csv,
)
from cajuru_a1.matcher import match_all
from cajuru_a1.models import JetaxClient, PipelineResult
from cajuru_a1.pfx import inspect_file
from cajuru_a1.passwords import PasswordVault, candidate_passwords

from conftest import make_pfx


def _client(cnpj, nome="EMPRESA"):
    return JetaxClient(razao_social=nome, cnpj=cnpj)


def _open(path, cnpj, empresa, senha):
    vault = PasswordVault()
    cands = candidate_passwords(vault=vault, empresa=empresa, cnpj=cnpj,
                                extra_names=[empresa], include_empty=True) + [(senha, "t")]
    return inspect_file(path, path, cands)


def test_lote_manual_senha_em_branco_e_csv(tmp_path):
    (tmp_path / "src").mkdir()
    p = tmp_path / "src" / "c.pfx"
    make_pfx(p, password="segredo123", cnpj="12345678000195", company="ABC LTDA")
    cert = _open(p, "12345678000195", "ABC", "segredo123")
    client = _client("12345678000195", "ABC LTDA")
    matches = match_all([cert], [client], [])
    pronto = [m for m in matches if m.status == "pronto"]
    assert pronto

    # Planilha com senha em branco
    zip_path, plan_path = build_importacao_jettax(pronto, tmp_path / "b1", senha_manual=True)
    assert zip_path.is_file() and plan_path.is_file()
    import openpyxl
    wb = openpyxl.load_workbook(plan_path)
    ws = wb["Certificados"]
    assert ws.cell(2, 2).value in (None, "")  # COL_SENHA=2 em branco
    assert ws.cell(2, 1).value  # CNPJ preenchido
    wb.close()

    # CSV de senhas separado
    csv_path = build_senhas_csv(pronto, tmp_path / "b1" / "senhas.csv")
    content = csv_path.read_text(encoding="utf-8-sig")
    assert "segredo123" in content  # CSV contém a senha para preenchimento manual
    assert "12.345.678/0001-95" in content


def test_lote_persistente(tmp_path):
    (tmp_path / "src").mkdir()
    p = tmp_path / "src" / "c.pfx"
    make_pfx(p, password="pw", cnpj="12345678000195", company="ABC")
    cert = _open(p, "12345678000195", "ABC", "pw")
    client = _client("12345678000195", "ABC")
    matches = match_all([cert], [client], [])
    pronto = [m for m in matches if m.status == "pronto"]

    out = tmp_path / "output"
    bundle = build_persistent_bundle(pronto, out, senha_manual=True, salvar_senhas_csv=True)
    assert bundle["dir"].is_dir()
    assert bundle["zip"].is_file()
    assert bundle["planilha"].is_file()
    assert bundle["csv_senhas"] and bundle["csv_senhas"].is_file()
    assert bundle["leia_me"].is_file()
    # não foi apagado
    assert bundle["dir"].exists()


def test_diagnostico_conteudo(tmp_path):
    (tmp_path / "src").mkdir()
    p_ok = tmp_path / "src" / "ok.pfx"
    p_venc = tmp_path / "src" / "venc.pfx"
    make_pfx(p_ok, password="pw", cnpj="12345678000195", company="OK", days_valid=300)
    make_pfx(p_venc, password="pw", cnpj="99999999000191", company="VELHA", days_valid=-5)

    c_ok = _open(p_ok, "12345678000195", "OK", "pw")
    c_venc = _open(p_venc, "99999999000191", "VELHA", "pw")
    clients = [_client("12345678000195", "OK"), _client("99999999000191", "VELHA")]
    matches = match_all([c_ok, c_venc], clients, [])

    result = PipelineResult(
        certificados=[c_ok, c_venc], clientes_sem=clients, clientes_com=[],
        matches=matches, temp_dir="", stats={}, safety_ok=True,
        source_manifest={}, source_inventory={}, source_root=str(tmp_path),
    )
    linhas = build_diagnostico(result)
    by_status = {linha.status: linha for linha in linhas}
    assert "pronto" in by_status
    assert "vencido" in by_status
    linha_venc = by_status["vencido"]
    assert linha_venc.dias is not None and linha_venc.dias < 0
    assert "VENCIDO" in linha_venc.situacao_validade

    xlsx = write_diagnostico_excel(linhas, tmp_path / "diag.xlsx")
    html = write_diagnostico_html(linhas, tmp_path / "diag.html")
    assert xlsx.is_file() and html.is_file()
    assert "Diagnóstico" in html.read_text(encoding="utf-8")


def test_pfx_sem_senha_abre(tmp_path):
    (tmp_path / "src").mkdir()
    p = tmp_path / "src" / "np.pfx"
    make_pfx(p, password=None, cnpj="12345678000195", company="NP")
    vault = PasswordVault()
    cands = candidate_passwords(vault=vault, empresa="NP", cnpj="12345678000195",
                                extra_names=["NP"], include_empty=True)
    info = inspect_file(p, p, cands)
    assert info.opened
    assert info.password in ("", None)


def test_list_all_clients_detecta_filtro_nao_aplicado(monkeypatch):
    """Se 'sem certificado' e 'com certificado' vierem com o mesmo conjunto de
    CNPJs, o filtro quase certamente não pegou de verdade em uma das duas
    telas do Jettax — list_all_clients() deve travar em vez de devolver dados
    errados silenciosamente para o matcher."""
    from cajuru_a1.jettax import JettaxBot

    bot = JettaxBot({"jettax": {"url": "https://admin.jettax360.com.br"}})
    mesma_lista = [_client(f"1234567800019{i}", f"EMPRESA {i}") for i in range(5)]
    monkeypatch.setattr(bot, "list_without_certificate", lambda: list(mesma_lista))
    monkeypatch.setattr(bot, "list_with_certificate", lambda: list(mesma_lista))

    with pytest.raises(RuntimeError, match="quase idênticas"):
        bot.list_all_clients()


def test_list_all_clients_ok_quando_listas_diferentes(monkeypatch):
    """Listas genuinamente diferentes (o caso normal) não devem disparar a
    checagem de segurança."""
    from cajuru_a1.jettax import JettaxBot

    bot = JettaxBot({"jettax": {"url": "https://admin.jettax360.com.br"}})
    sem = [_client(f"1111111100019{i}", f"SEM {i}") for i in range(5)]
    com = [_client(f"2222222200019{i}", f"COM {i}") for i in range(5)]
    monkeypatch.setattr(bot, "list_without_certificate", lambda: list(sem))
    monkeypatch.setattr(bot, "list_with_certificate", lambda: list(com))

    got_sem, got_com = bot.list_all_clients()
    assert got_sem == sem
    assert got_com == com
