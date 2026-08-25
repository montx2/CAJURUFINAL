from __future__ import annotations

import csv
import os
import shutil
import stat
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl

from cajuru_a1.cnpjutil import format_cnpj, is_valid_doc, only_digits, pad_cnpj

# Modelo oficial de importação em lote do Jettax 360 (aba "Certificados").
# Baixado do próprio Jettax; nunca editar os cabeçalhos/ordem das colunas,
# só as linhas de dados (a partir da linha 2).
TEMPLATE_PATH = Path(__file__).parent / "resources" / "modelo_import_certificados.xlsx"
TEMPLATE_SHEET = "Certificados"
# Colunas obrigatórias que este programa preenche. As demais (regime
# tributário, credenciais de prefeitura, módulos, etc.) só importam para
# cliente NOVO — os nossos são todos clientes já cadastrados sem A1, então
# ficam em branco de propósito.
COL_CNPJ = 1
COL_SENHA = 2


def _clear_datalines(sheet) -> None:
    """Apaga valores e hiperlinks das linhas de dados (a partir da linha 2),
    preservando o cabeçalho e a formatação das colunas do modelo oficial."""
    for row in range(2, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, col)
            cell.value = None
            try:
                cell.hyperlink = None
            except (AttributeError, TypeError):
                pass


def _within(path: Path, root: Path) -> bool:
    try:
        path.resolve(strict=False).relative_to(root.resolve(strict=False))
        return True
    except ValueError:
        return False


def _guard_destino(destination: Path, matches) -> None:
    if any(part.casefold() == "dropbox" or part.casefold().startswith("dropbox (") for part in destination.parts):
        raise ValueError("Lote não pode ser criado dentro de uma árvore Dropbox")
    for match in matches:
        source_path = getattr(getattr(match, "cert", None), "source_path", "")
        if source_path and _within(destination, Path(source_path).parent):
            raise ValueError("Lote não pode ser criado dentro da origem Dropbox")


def _elegiveis(matches) -> list:
    elegiveis = []
    documentos: set[str] = set()
    for match in matches:
        cert = getattr(match, "cert", None)
        client = getattr(match, "cliente", None)
        if not getattr(match, "pode_enviar", False) or not cert or not client or not cert.temp_path:
            continue
        document = pad_cnpj(only_digits(client.cnpj))
        if not is_valid_doc(document):
            raise RuntimeError("Documento inválido no lote")
        if document in documentos:
            # A própria planilha do Jettax invalida as duas linhas em caso de
            # CNPJ repetido; melhor barrar aqui e obrigar revisão manual.
            raise RuntimeError(f"CNPJ duplicado no lote ({format_cnpj(document)}); envio bloqueado")
        documentos.add(document)
        extension = Path(cert.temp_path).suffix.casefold()
        if extension not in {".pfx", ".p12"}:
            raise RuntimeError("Extensão de certificado inválida no lote")
        elegiveis.append((match, document))
    return elegiveis


def build_certificados_zip(matches, dest_dir: Path) -> Path:
    """ZIP só com os .pfx, cada um nomeado exatamente como o CNPJ (só dígitos),
    conforme exigido pelo importador do Jettax (ex.: 12345678000195.pfx).
    Artefato transitório; o chamador deve apagá-lo em finally.
    """
    destination = Path(dest_dir).expanduser().resolve(strict=False)
    elegiveis = _elegiveis(matches)
    _guard_destino(destination, [m for m, _ in elegiveis])
    destination.mkdir(parents=True, exist_ok=True)
    try:
        destination.chmod(stat.S_IRWXU)
    except OSError:
        pass
    work = destination / "_certificados_jettax"
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(mode=0o700)

    if not elegiveis:
        raise RuntimeError("Nenhum certificado PRONTO e elegível para o lote")

    for match, document in elegiveis:
        cert = match.cert
        source = Path(cert.temp_path)
        if not source.is_file():
            raise FileNotFoundError(source)
        # O Jettax só documenta a extensão .pfx; PKCS#12 (.p12) tem o mesmo
        # formato binário, então renomear para .pfx é seguro e não altera o
        # conteúdo do certificado.
        target = work / f"{document}.pfx"
        with source.open("rb") as input_stream, target.open("xb") as output_stream:
            shutil.copyfileobj(input_stream, output_stream, length=1024 * 1024)
        try:
            target.chmod(stat.S_IRUSR | stat.S_IWUSR)
        except OSError:
            pass

    output = destination / "certificados_jettax.zip"
    temporary = destination / ".certificados_jettax.zip.tmp"
    with zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as archive:
        for file in sorted(work.iterdir()):
            if file.is_file():
                archive.write(file, file.name)
    os.replace(temporary, output)
    try:
        output.chmod(stat.S_IRUSR | stat.S_IWUSR)
    except OSError:
        pass
    shutil.rmtree(work, ignore_errors=True)
    return output


def build_planilha_importacao(
    matches,
    dest_dir: Path,
    template_path: Path | None = None,
    *,
    senha_manual: bool = False,
) -> Path:
    """Preenche o modelo OFICIAL do Jettax (aba 'Certificados') com CNPJ e
    Senha Certificado de cada match elegível.

    Quando ``senha_manual=True``, a coluna de senha fica em branco para o
    usuário preencher à mão (modo mais seguro — o programa não escreve a
    senha do A1 na planilha de importação). Quando False, o comportamento
    histórico é mantido: a senha validada é preenchida automaticamente.
    """
    template_path = Path(template_path) if template_path else TEMPLATE_PATH
    if not template_path.is_file():
        raise FileNotFoundError(
            f"Modelo oficial de importação não encontrado em {template_path}. "
            "Baixe o modelo mais recente no próprio Jettax (botão Importar) e "
            "salve em cajuru_a1/resources/modelo_import_certificados.xlsx."
        )
    destination = Path(dest_dir).expanduser().resolve(strict=False)
    elegiveis = _elegiveis(matches)
    _guard_destino(destination, [m for m, _ in elegiveis])
    if not elegiveis:
        raise RuntimeError("Nenhum certificado PRONTO e elegível para o lote")
    destination.mkdir(parents=True, exist_ok=True)
    try:
        destination.chmod(stat.S_IRWXU)
    except OSError:
        pass

    workbook = openpyxl.load_workbook(template_path)
    if TEMPLATE_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"O modelo não tem a aba obrigatória '{TEMPLATE_SHEET}'")
    sheet = workbook[TEMPLATE_SHEET]
    _clear_datalines(sheet)

    for index, (match, document) in enumerate(elegiveis, start=2):
        cert = match.cert
        # CNPJ formatado com máscara, igual ao exemplo do próprio modelo
        # ("AA.000.000/0000-00"). Força texto para não perder zero à esquerda
        # nem virar fórmula.
        cnpj_cell = sheet.cell(index, COL_CNPJ)
        cnpj_cell.value = format_cnpj(document)
        cnpj_cell.data_type = "s"
        senha_cell = sheet.cell(index, COL_SENHA)
        if senha_manual:
            # Modo manual: o usuário digita a senha do A1 diretamente no
            # Jettax / na planilha. Não persistimos a senha.
            senha_cell.value = ""
        else:
            senha_cell.value = "" if cert.password is None else str(cert.password)
        senha_cell.data_type = "s"

    output = destination / "planilha_importacao_jettax.xlsx"
    temporary = destination / ".planilha_importacao_jettax.xlsx.tmp"
    workbook.save(temporary)
    workbook.close()
    os.replace(temporary, output)
    try:
        output.chmod(stat.S_IRUSR | stat.S_IWUSR)
    except OSError:
        pass
    return output


def build_senhas_csv(matches, dest_path: Path) -> Path | None:
    """Gera um CSV separado com as senhas validadas (CNPJ, empresa, arquivo, senha).

    Este arquivo é o 'meio de campo' para o modo manual: em vez de escrever a
    senha na planilha oficial do Jettax, deixamos o ZIP + planilha oficial com
    a coluna de senha em branco e geramos este CSV para o usuário consultar e
    digitar/copiar manualmente. É óbvio que contém segredos, por isso fica
    claramente nomeado e é salvo fora do Dropbox.
    """
    elegiveis = _elegiveis(matches)
    if not elegiveis:
        return None
    dest_path = Path(dest_path).expanduser().resolve(strict=False)
    _guard_destino(dest_path.parent, [m for m, _ in elegiveis])
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    with dest_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["CNPJ", "EMPRESA", "ARQUIVO_PFX", "SENHA_CERTIFICADO", "VALIDADE"])
        for match, document in elegiveis:
            cert = match.cert
            client = match.cliente
            validade = cert.not_after.strftime("%d/%m/%Y") if getattr(cert, "not_after", None) else ""
            w.writerow([
                format_cnpj(document),
                client.razao_social if client else "",
                cert.filename,
                "" if cert.password is None else str(cert.password),
                validade,
            ])
    try:
        dest_path.chmod(stat.S_IRUSR | stat.S_IWUSR)
    except OSError:
        pass
    return dest_path


def build_planilha_importacao_certificados(
    certificados,
    dest_dir: Path,
    template_path: Path | None = None,
    *,
    senha_manual: bool = False,
) -> Path:
    """Preenche o modelo OFICIAL do Jettax (aba 'Certificados') a partir de uma
    lista de certificados A1 já abertos (objetos com ``cnpj``/``password``).

    É a versão usada pela exportação local de "todos os certificados + senha"
    (sem conciliação nem login no Jettax): preenche CNPJ e SENHA de cada
    certificado válido, seguindo exatamente o modelo
    ``modelo_import_certificados.xlsx`` (cabeçalhos, abas e ordem das colunas
    preservados). Apenas certificados já abertos e com documento válido são
    escritos; os demais continuam disponíveis somente no ZIP/CSV.
    """
    template_path = Path(template_path) if template_path else TEMPLATE_PATH
    if not template_path.is_file():
        raise FileNotFoundError(
            f"Modelo oficial de importação não encontrado em {template_path}. "
            "Baixe o modelo mais recente no próprio Jettax (botão Importar) e "
            "salve em cajuru_a1/resources/modelo_import_certificados.xlsx."
        )
    destination = Path(dest_dir).expanduser().resolve(strict=False)
    destination.mkdir(parents=True, exist_ok=True)
    try:
        destination.chmod(stat.S_IRWXU)
    except OSError:
        pass

    workbook = openpyxl.load_workbook(template_path)
    if TEMPLATE_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"O modelo não tem a aba obrigatória '{TEMPLATE_SHEET}'")
    sheet = workbook[TEMPLATE_SHEET]
    _clear_datalines(sheet)

    index = 2
    written = 0
    for cert in certificados or []:
        if not getattr(cert, "opened", False):
            continue
        document = pad_cnpj(only_digits(getattr(cert, "cnpj", "") or ""))
        if not is_valid_doc(document):
            continue
        cnpj_cell = sheet.cell(index, COL_CNPJ)
        cnpj_cell.value = format_cnpj(document)
        cnpj_cell.data_type = "s"
        senha_cell = sheet.cell(index, COL_SENHA)
        # Senha validada preenchida por padrão (modo exportação), ou em branco
        # se o chamador pedir senha manual.
        senha_cell.value = "" if senha_manual or cert.password is None else str(cert.password)
        senha_cell.data_type = "s"
        index += 1
        written += 1
    if not written:
        raise RuntimeError("Nenhum certificado aberto e com CNPJ válido para preencher a planilha de importação.")

    output = destination / "planilha_importacao_jettax.xlsx"
    temporary = destination / ".planilha_importacao_jettax.xlsx.tmp"
    workbook.save(temporary)
    workbook.close()
    os.replace(temporary, output)
    try:
        output.chmod(stat.S_IRUSR | stat.S_IWUSR)
    except OSError:
        pass
    return output


def build_importacao_jettax(
    matches,
    dest_dir: Path,
    template_path: Path | None = None,
    *,
    senha_manual: bool = False,
) -> tuple[Path, Path]:
    """Gera os dois artefatos exigidos pelo importador oficial do Jettax:
    o ZIP com os .pfx (nomeados por CNPJ) e a planilha preenchida a partir
    do modelo oficial. Devem ser enviados juntos na tela 'Importar' do Jettax.
    """
    zip_path = build_certificados_zip(matches, dest_dir)
    planilha_path = build_planilha_importacao(
        matches, dest_dir, template_path=template_path, senha_manual=senha_manual
    )
    return zip_path, planilha_path


def build_persistent_bundle(
    matches,
    output_dir: Path,
    *,
    template_path: Path | None = None,
    senha_manual: bool = True,
    salvar_senhas_csv: bool = True,
) -> dict:
    """Salva o ZIP + planilha (e opcionalmente o CSV de senhas) em
    ``<output_dir>/lotes/lote_YYYYmmdd_HHMMSS/`` e devolve os caminhos.

    Diferente do lote transitório do envio automático, estes arquivos NÃO são
    apagos em seguida — é o modo de operação 100% manual solicitado: o usuário
    leva os dois arquivos ao Jettax e digita a senha ele mesmo.
    """
    output_dir = Path(output_dir).expanduser().resolve(strict=False)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    bundle_dir = output_dir / "lotes" / f"lote_{stamp}"
    bundle_dir.mkdir(parents=True, exist_ok=True)
    zip_path, planilha_path = build_importacao_jettax(
        matches, bundle_dir, template_path=template_path, senha_manual=senha_manual
    )
    csv_path: Path | None = None
    if salvar_senhas_csv:
        csv_path = build_senhas_csv(matches, bundle_dir / "senhas_para_preenchimento_manual.csv")
    leia_me = bundle_dir / "LEIA-ME.txt"
    leia_me.write_text(
        "LOTE DE IMPORTACAO JETTAX 360 (modo manual)\n"
        "============================================\n\n"
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n"
        "Arquivos para importar NO JETTAX:\n"
        f"  1. {zip_path.name}  (ZIP com os .pfx nomeados por CNPJ)\n"
        f"  2. {planilha_path.name}  (planilha oficial preenchida; coluna SENHA em branco)\n\n"
        + (
            f"  3. {csv_path.name}  (senhas validadas — CONSULTE E DIGITE MANUALMENTE; apague depois)\n\n"
            if csv_path else "\n"
        )
        + "No Jettax:\n"
        "  - Clientes > Importar\n"
        "  - Selecione o ZIP e a planilha\n"
        "  - Confira cada CNPJ e digite a senha do A1 (veja o CSV de senhas)\n"
        "  - Conclua a importacao\n\n"
        "AVISO: estes arquivos ficam salvos aqui propositalmente (modo manual).\n"
        "Eles contem senhas de certificados. Depois de importar, apague esta pasta\n"
        "manual ou mova para local seguro. O Dropbox nunca foi alterado.\n",
        encoding="utf-8",
    )
    return {
        "dir": bundle_dir,
        "zip": zip_path,
        "planilha": planilha_path,
        "csv_senhas": csv_path,
        "leia_me": leia_me,
        "senha_manual": senha_manual,
    }


def cleanup_import_bundle(*paths: Path) -> None:
    """Apaga com segurança os artefatos transitórios com senha (ZIP, planilha
    e a pasta-mãe exclusiva criada para o lote)."""
    errors: list[Exception] = []
    parents: set[Path] = set()
    for raw in paths:
        if not raw:
            continue
        path = Path(raw).expanduser().resolve(strict=False)
        parents.add(path.parent)
        for candidate in (
            path,
            path.parent / "_certificados_jettax",
            path.parent / ".certificados_jettax.zip.tmp",
            path.parent / ".planilha_importacao_jettax.xlsx.tmp",
        ):
            try:
                if candidate.is_dir():
                    shutil.rmtree(candidate)
                elif candidate.exists():
                    candidate.unlink(missing_ok=True)
            except OSError as exc:
                errors.append(exc)
    for parent in parents:
        if parent.name.startswith("cajuru_a1_bundle_"):
            try:
                parent.rmdir()
            except OSError as exc:
                errors.append(exc)
    if errors:
        raise RuntimeError("Não foi possível remover completamente o lote transitório com senhas") from errors[0]
