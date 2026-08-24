"""Relatório de diagnóstico: por que não funcionou, validade e histórico.

Reúne, para cada certificado:
- motivo da falha / bloqueio (código + mensagem legível);
- datas de validade (início, fim, dias para vencer ou dias vencido);
- tentativas de senha;
- estado em execuções anteriores (o que tinha antes);
- a que cliente Jettax ele se refere (se houver).

Gera um HTML rico e um Excel (aba "Diagnóstico") sem valor de senha.
"""

from __future__ import annotations

import html
import os
import tempfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from cajuru_a1.cnpjutil import format_cnpj
from cajuru_a1.models import PipelineResult

ERROR_LABELS = {
    "": "OK",
    "arquivo_grande": "Arquivo excede o limite de segurança",
    "pfx_corrompido": "PFX vazio ou corrompido (não é PKCS#12 válido)",
    "senha_nao_encontrada_ou_pfx_invalido": "Nenhuma senha plausível abriu (ou PFX inválido)",
    "falha_copia": "Falha de cópia/integridade ao sair do Dropbox",
    "falha_inspecao": "Falha isolada durante a inspeção",
    "timeout_pfx": "PKCS#12 travou o OpenSSL e foi interrompido (revisão manual)",
    "origem_insegura": "Origem insegura/ilegível bloqueada (symlink etc.)",
    "identidade_interna_ambigua": "O certificado contém mais de um documento",
    "cnpj_nome_diferente_certificado": "CNPJ do nome do arquivo é diferente do CNPJ interno",
    "pdf_grande": "PDF excede o limite de segurança",
    "pdf_corrompido": "PDF ilegível",
    "pdf_senha_incorreta": "PDF protegido e nenhuma senha funcionou",
}

STATUS_LABEL = {
    "pronto": "PRONTO",
    "vencido": "VENCIDO",
    "nao_valido": "AINDA NÃO VÁLIDO",
    "sem_senha": "SEM SENHA",
    "invalido": "INVÁLIDO (sem chave)",
    "conflito": "CONFLITO",
    "ambiguo": "AMBÍGUO",
    "revisao_manual": "REVISÃO MANUAL",
    "substituido": "SUBSTITUÍDO POR MAIS NOVO",
    "duplicado": "DUPLICADO",
    "sem_cert": "CLIENTE SEM PFX",
    "sem_cert_novo": "CLIENTE JÁ TEM A1 (sem PFX novo)",
    "extra_pfx": "PFX SEM CLIENTE",
}

STATUS_COLOR = {
    "pronto": "1B9C85",
    "vencido": "C0392B",
    "nao_valido": "8E6BBE",
    "sem_senha": "E0A100",
    "invalido": "922B21",
    "conflito": "D35400",
    "ambiguo": "8E44AD",
    "revisao_manual": "AF7AC5",
    "substituido": "566573",
    "duplicado": "7F8C8D",
    "sem_cert": "7F8C8D",
    "sem_cert_novo": "2E86C1",
    "extra_pfx": "2980B9",
}


@dataclass
class DiagnosticoLinha:
    status: str
    empresa: str = ""
    cnpj: str = ""
    arquivo: str = ""
    cnpj_interno: str = ""
    aberto: bool = False
    validade_inicio: str = ""
    validade_fim: str = ""
    dias: int | None = None
    situacao_validade: str = ""
    tentativas_senha: int = 0
    origem_senha: str = ""
    codigo_erro: str = ""
    motivo: str = ""
    caminho_dropbox: str = ""
    sha256: str = ""
    historico: list[dict[str, Any]] = field(default_factory=list)

    def to_row(self) -> list[Any]:
        hist = self.historico[0] if self.historico else {}
        hist_texto = ""
        if hist:
            hist_texto = (
                f"{hist.get('status', '')} em {hist.get('updated_utc', '')[:19]}; "
                f"erro={hist.get('error_code', '') or '-'}; tentativas={hist.get('attempts', 0)}"
            )
        return [
            STATUS_LABEL.get(self.status, self.status.upper()),
            self.empresa,
            format_cnpj(self.cnpj) if self.cnpj else "",
            self.arquivo,
            format_cnpj(self.cnpj_interno) if self.cnpj_interno else "",
            "SIM" if self.aberto else "NÃO",
            self.validade_inicio,
            self.validade_fim,
            self.dias if self.dias is not None else "",
            self.situacao_validade,
            self.tentativas_senha,
            self.origem_senha,
            ERROR_LABELS.get(self.codigo_erro, self.codigo_erro),
            self.motivo,
            hist_texto,
            self.caminho_dropbox,
            self.sha256[:16] if self.sha256 else "",
        ]


def _validade_situation(cert) -> tuple[str, str, int | None]:
    if not getattr(cert, "opened", False):
        return ("", "", None)
    now = datetime.now(timezone.utc)
    if getattr(cert, "not_after", None) is None:
        return ("desconhecida", "", None)
    dias = (cert.not_after - now).days
    if getattr(cert, "expired", False):
        return (f"VENCIDO há {abs(dias)} dia(s)", cert.not_after.strftime("%d/%m/%Y"), dias)
    if getattr(cert, "not_yet_valid", False):
        return ("AINDA NÃO VÁLIDO", getattr(cert, "not_before", cert.not_after).strftime("%d/%m/%Y"), dias)
    if dias <= 30:
        return (f"VENCE EM {dias} dia(s)", cert.not_after.strftime("%d/%m/%Y"), dias)
    return (f"válido por {dias} dia(s)", cert.not_after.strftime("%d/%m/%Y"), dias)


def build_diagnostico(result: PipelineResult, state=None) -> list[DiagnosticoLinha]:
    linhas: list[DiagnosticoLinha] = []
    # Um PFX pode aparecer em mais de um match? Não — o matcher usa cada PFX uma
    # única vez. Mesmo assim indexamos por id(cert) para não duplicar.
    seen_certs: set[int] = set()
    for match in result.matches:
        cert = match.cert
        client = match.cliente
        if cert is not None and id(cert) in seen_certs:
            continue
        if cert is not None:
            seen_certs.add(id(cert))
        linha = DiagnosticoLinha(
            status=match.status,
            empresa=(client.razao_social if client else "") or (getattr(cert, "company_from_cert", "") if cert else ""),
            cnpj=(client.cnpj if client else "") or "",
            arquivo=getattr(cert, "filename", "") if cert else "",
            cnpj_interno=getattr(cert, "cnpj_cert", "") or "",
            aberto=bool(getattr(cert, "opened", False)) if cert else False,
            validade_inicio=(
                cert.not_before.strftime("%d/%m/%Y")
                if cert and getattr(cert, "not_before", None) else ""
            ),
            codigo_erro=getattr(cert, "error_code", "") if cert else "",
            motivo=match.motivo or (getattr(cert, "error", "") if cert else ""),
            caminho_dropbox=getattr(cert, "source_path", "") if cert else "",
            sha256=getattr(cert, "sha256", "") if cert else "",
            tentativas_senha=int(getattr(cert, "attempts", 0) or 0) if cert else 0,
            origem_senha=getattr(cert, "password_source", "") or "" if cert else "",
        )
        if cert:
            sit, fim, dias = _validade_situation(cert)
            linha.situacao_validade = sit
            linha.validade_fim = fim
            linha.dias = dias
        if state is not None and cert and getattr(cert, "source_path", ""):
            try:
                from pathlib import Path as _P
                rel = _P(cert.source_path)
                # tenta caminho relativo; se falhar, usa nome
                root = getattr(result, "source_root", "")
                if root:
                    try:
                        rel = rel.relative_to(root)
                    except ValueError:
                        rel = _P(cert.filename)
                linha.historico = state.file_history(rel.as_posix())
            except Exception:
                linha.historico = []
        linhas.append(linha)
    return linhas


def write_diagnostico_excel(linhas: list[DiagnosticoLinha], dest: Path) -> Path:
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Diagnóstico"
    headers = [
        "STATUS", "EMPRESA", "CNPJ", "ARQUIVO PFX", "CNPJ INTERNO", "ABERTO",
        "INÍCIO VALIDADE", "FIM VALIDADE", "DIAS", "SITUAÇÃO",
        "TENTATIVAS SENHA", "ORIGEM SENHA", "ERRO", "MOTIVO", "HISTÓRICO ANTERIOR",
        "CAMINHO DROPBOX", "SHA-256 (16)",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B1F3A")
    thin = Border(**{s: Side(style="thin", color="DDDDDD") for s in ("left", "right", "top", "bottom")})
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for linha in linhas:
        ws.append(linha.to_row())
        color = STATUS_COLOR.get(linha.status, "808080")
        c = ws.cell(ws.max_row, 1)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, color="FFFFFF")
        for col in range(1, len(headers) + 1):
            ws.cell(ws.max_row, col).border = thin
            ws.cell(ws.max_row, col).alignment = Alignment(vertical="top", wrap_text=True)
    widths = [18, 38, 20, 38, 20, 8, 14, 14, 8, 24, 12, 24, 40, 55, 55, 50, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes = "A2"

    # Aba de resumo
    resumo = wb.create_sheet("Resumo")
    resumo.append(["Cajuru A1 — Diagnóstico completo"])
    resumo.append(["Gerado em", datetime.now().astimezone().strftime("%d/%m/%Y %H:%M:%S")])
    resumo.append([])
    resumo.append(["Status", "Qtd"])
    counts: dict[str, int] = {}
    for linha in linhas:
        counts[linha.status] = counts.get(linha.status, 0) + 1
    for status, count in sorted(counts.items(), key=lambda x: -x[1]):
        resumo.append([STATUS_LABEL.get(status, status), count])
    vencidos = sum(1 for linha in linhas if linha.status == "vencido" or (linha.dias is not None and linha.dias < 0))
    a_vencer = sum(1 for linha in linhas if linha.dias is not None and 0 <= linha.dias <= 30 and linha.aberto)
    resumo.append([])
    resumo.append(["Total de certificados", len(linhas)])
    resumo.append(["Vencidos", vencidos])
    resumo.append(["A vencer em 30 dias", a_vencer])
    resumo.column_dimensions["A"].width = 36
    resumo.column_dimensions["B"].width = 18

    fd, tmp = tempfile.mkstemp(prefix=".diagnostico.", suffix=".xlsx", dir=str(dest.parent))
    os.close(fd)
    try:
        wb.save(tmp)
        wb.close()
        os.replace(tmp, dest)
    finally:
        Path(tmp).unlink(missing_ok=True)
    return dest


def write_diagnostico_html(linhas: list[DiagnosticoLinha], dest: Path, *, stats: dict | None = None) -> Path:
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)

    def esc(v) -> str:
        return html.escape("" if v is None else str(v))

    counts: dict[str, int] = {}
    for linha in linhas:
        counts[linha.status] = counts.get(linha.status, 0) + 1
    cards = "".join(
        f"<div class='card {s}'><b>{n}</b><span>{esc(STATUS_LABEL.get(s, s))}</span></div>"
        for s, n in sorted(counts.items(), key=lambda x: -x[1])
    )
    rows_html = []
    for linha in linhas:
        color = "#" + STATUS_COLOR.get(linha.status, "808080")
        hist = linha.historico[0] if linha.historico else {}
        hist_txt = (
            f"{hist.get('status','')} · {str(hist.get('updated_utc',''))[:19]} · "
            f"erro={hist.get('error_code','') or '—'} · tentativas={hist.get('attempts',0)}"
            if hist else "—"
        )
        rows_html.append(
            "<tr>"
            f"<td><span class='badge' style='background:{color}'>{esc(STATUS_LABEL.get(linha.status, linha.status))}</span></td>"
            f"<td>{esc(linha.empresa)}</td>"
            f"<td>{esc(format_cnpj(linha.cnpj) if linha.cnpj else '')}</td>"
            f"<td class='mono'>{esc(linha.arquivo)}</td>"
            f"<td>{esc(format_cnpj(linha.cnpj_interno) if linha.cnpj_interno else '')}</td>"
            f"<td class='center'>{'✓' if linha.aberto else '✗'}</td>"
            f"<td>{esc(linha.validade_inicio)}</td>"
            f"<td>{esc(linha.validade_fim)}</td>"
            f"<td class='center'>{esc(linha.dias if linha.dias is not None else '')}</td>"
            f"<td>{esc(linha.situacao_validade)}</td>"
            f"<td class='center'>{linha.tentativas_senha}</td>"
            f"<td>{esc(linha.origem_senha)}</td>"
            f"<td>{esc(ERROR_LABELS.get(linha.codigo_erro, linha.codigo_erro))}</td>"
            f"<td>{esc(linha.motivo)}</td>"
            f"<td class='mono small'>{esc(hist_txt)}</td>"
            "</tr>"
        )
    total = len(linhas)
    vencidos = sum(1 for linha in linhas if linha.status == "vencido" or (linha.dias is not None and linha.dias < 0))
    a_vencer = sum(1 for linha in linhas if linha.dias is not None and 0 <= linha.dias <= 30 and linha.aberto)
    document = f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Cajuru A1 — Diagnóstico completo</title>
<style>
*{{box-sizing:border-box}}
body{{margin:0;font-family:'Segoe UI',Roboto,Arial,sans-serif;background:#0A0C11;color:#EBEDF1}}
header{{background:linear-gradient(135deg,#12151C,#1B2340);padding:28px 40px;border-bottom:1px solid #242A36}}
h1{{margin:0 0 4px;font-size:22px;font-weight:650;letter-spacing:-.01em}} .sub{{color:#8890A0;font-size:13px}}
.wrap{{padding:24px 40px 60px;max-width:1600px;margin:0 auto}}
.cards{{display:flex;gap:12px;flex-wrap:wrap;margin:18px 0 24px}}
.card{{background:#12151C;border-radius:12px;padding:16px 20px;min-width:170px;border:1px solid #242A36}}
.card b{{display:block;font-size:26px;font-weight:650}} .card span{{color:#8890A0;font-size:11.5px;text-transform:uppercase;letter-spacing:.5px}}
.card.vencido b,.card.invalido b,.card.conflito b{{color:#DA5257}} .card.sem_senha b,.card.ambiguo b{{color:#D3941F}}
.card.pronto b,.card.sem_cert_novo b{{color:#2FA968}}
.kpis{{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:22px}}
.kpi{{background:#171B24;border:1px solid #242A36;border-radius:12px;padding:14px 20px;flex:1;min-width:180px}}
.kpi b{{font-size:22px;display:block;font-weight:650}} .kpi span{{color:#8890A0;font-size:11.5px}}
table{{width:100%;border-collapse:collapse;background:#12151C;border:1px solid #242A36;border-radius:12px;overflow:hidden;font-size:13px}}
th{{background:#171B24;color:#5B6273;text-align:left;padding:11px 10px;position:sticky;top:0;font-weight:600;font-size:10.5px;text-transform:uppercase;letter-spacing:.5px}}
td{{padding:9px 10px;border-bottom:1px solid #1A1F29;vertical-align:top}}
tr:hover td{{background:#1B2029}}
.badge{{color:#fff;padding:3px 10px;border-radius:999px;font-size:11px;text-transform:uppercase;white-space:nowrap;display:inline-block}}
.mono{{font-family:'JetBrains Mono',Consolas,'SF Mono',monospace;font-size:12px;color:#EBEDF1}}
.small{{font-size:11px;color:#5B6273}} .center{{text-align:center}}
.note{{margin-top:18px;color:#5B6273;font-size:12px;line-height:1.6}}
.toolbar{{margin-bottom:14px;display:flex;gap:10px;align-items:center;flex-wrap:wrap}}
input[type=search]{{background:#171B24;border:1px solid #242A36;color:#EBEDF1;border-radius:8px;padding:9px 12px;width:320px;font-size:13px}}
select{{background:#171B24;border:1px solid #242A36;color:#EBEDF1;border-radius:8px;padding:9px 12px;font-size:13px}}
</style></head><body>
<header><h1>Diagnóstico completo dos certificados</h1>
<div class="sub">Gerado em {datetime.now().astimezone().strftime('%d/%m/%Y %H:%M')} · {total} certificado(s) · Nenhum valor de senha é exibido</div></header>
<div class="wrap">
<div class="kpis">
  <div class="kpi"><b>{total}</b><span>Total de certificados</span></div>
  <div class="kpi"><b style="color:#DA5257">{vencidos}</b><span>Vencidos</span></div>
  <div class="kpi"><b style="color:#D3941F">{a_vencer}</b><span>A vencer em 30 dias</span></div>
  <div class="kpi"><b style="color:#2FA968">{counts.get('pronto',0)}</b><span>Prontos</span></div>
</div>
<div class="cards">{cards}</div>
<div class="toolbar">
  <input id="q" type="search" placeholder="Buscar por empresa, CNPJ ou arquivo..." oninput="filtrar()"/>
  <select id="st" onchange="filtrar()"><option value="">Todos os status</option>{''.join(f'<option>{esc(STATUS_LABEL.get(s,s))}</option>' for s in sorted(set(list(STATUS_LABEL.keys())+list(counts.keys()))) if s in counts)}</select>
</div>
<table id="tbl"><thead><tr>
<th>Status</th><th>Empresa</th><th>CNPJ</th><th>Arquivo</th><th>CNPJ interno</th><th>Aberto</th>
<th>Início</th><th>Fim</th><th>Dias</th><th>Situação</th><th>Tent.</th><th>Origem senha</th>
<th>Erro</th><th>Motivo</th><th>Histórico anterior</th></tr></thead><tbody>
{''.join(rows_html)}
</tbody></table>
<p class="note">A coluna <b>Histórico anterior</b> mostra o que o sistema registrou sobre o mesmo arquivo em execuções passadas
(status, data, código de erro e número de tentativas de senha). O Dropbox é origem somente leitura e não foi alterado.</p>
</div>
<script>
function filtrar(){{
  var q=document.getElementById('q').value.toLowerCase();
  var st=document.getElementById('st').value;
  document.querySelectorAll('#tbl tbody tr').forEach(function(tr){{
    var t=tr.innerText.toLowerCase();
    var okq=!q||t.indexOf(q)>=0;
    var badge=tr.querySelector('.badge');
    var oks=!st||(badge&&badge.innerText.trim()===st);
    tr.style.display=(okq&&oks)?'':'none';
  }});
}}
</script>
</body></html>"""
    fd, tmp = tempfile.mkstemp(prefix=".diagnostico.", suffix=".html", dir=str(dest.parent))
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as f:
            f.write(document)
        os.replace(tmp, dest)
    finally:
        Path(tmp).unlink(missing_ok=True)
    return dest
