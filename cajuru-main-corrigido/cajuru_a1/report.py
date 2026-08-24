from __future__ import annotations

import html
import os
import stat
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from cajuru_a1.cnpjutil import format_cnpj
from cajuru_a1.models import PipelineResult

COLORS = {
    "pronto": "1B9C85", "sem_senha": "E0A100", "vencido": "C0392B",
    "nao_valido": "8E6BBE", "invalido": "922B21", "duplicado": "7F8C8D",
    "sem_cert": "7F8C8D", "sem_cert_novo": "2E86C1", "ambiguo": "8E44AD",
    "conflito": "D35400", "revisao_manual": "AF7AC5", "extra_pfx": "2980B9",
    "substituido": "566573",
}


def _guard_destination(result: PipelineResult, destination: Path) -> Path:
    path = Path(destination).expanduser().resolve(strict=False)
    if any(part.casefold() == "dropbox" or part.casefold().startswith("dropbox (") for part in path.parts):
        raise ValueError("Relatório não pode ser gravado dentro de uma árvore Dropbox")
    if result.source_root:
        try:
            path.relative_to(Path(result.source_root).resolve(strict=False))
        except ValueError:
            pass
        else:
            raise ValueError("Relatório não pode ser gravado dentro da origem Dropbox")
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _excel_safe(value):
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@", "\t", "\r")):
        return "'" + value
    return value


def _append(sheet, values) -> None:
    sheet.append([_excel_safe(value) for value in values])


def write_excel_report(result: PipelineResult, dest: Path) -> Path:
    destination = _guard_destination(result, dest)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Conciliação"
    headers = [
        "STATUS", "EMPRESA JETTAX", "CNPJ", "ARQUIVO PFX", "CNPJ INTERNO",
        "SENHA VALIDADA", "ORIGEM DA CANDIDATA", "INÍCIO VALIDADE", "FIM VALIDADE",
        "DIAS P/ VENCER", "MÉTODO", "CONFIANÇA", "MOTIVO", "EVIDÊNCIAS",
        "ORIGEM DROPBOX (não alterada)", "HASH SHA-256",
    ]
    _append(sheet, headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B1F3A")
    thin = Border(**{side: Side(style="thin", color="DDDDDD") for side in ("left", "right", "top", "bottom")})
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for match in result.matches:
        client = match.cliente
        cert = match.cert
        dias = cert.dias_para_vencer if cert else None
        _append(sheet, [
            match.status.upper(),
            client.razao_social if client else "",
            format_cnpj(client.cnpj) if client and client.cnpj else "",
            cert.filename if cert else "",
            format_cnpj(cert.cnpj_cert) if cert and cert.cnpj_cert else "",
            "SIM" if cert and cert.opened else "NÃO",
            cert.password_source if cert else "",
            cert.not_before.strftime("%d/%m/%Y") if cert and cert.not_before else "",
            cert.not_after.strftime("%d/%m/%Y") if cert and cert.not_after else "",
            (dias if dias is not None else ""),
            match.metodo,
            round(match.confianca, 1) if match.confianca else "",
            match.motivo,
            " | ".join(match.evidencias),
            cert.source_path if cert else "",
            cert.sha256 if cert else "",
        ])
        color = PatternFill("solid", fgColor=COLORS.get(match.status, "FFFFFF"))
        sheet.cell(sheet.max_row, 1).fill = color
        sheet.cell(sheet.max_row, 1).font = Font(bold=True, color="FFFFFF")
        for column in range(1, len(headers) + 1):
            sheet.cell(sheet.max_row, column).border = thin

    widths = [17, 42, 20, 42, 20, 15, 31, 14, 14, 13, 22, 12, 52, 55, 50, 68]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.auto_filter.ref = f"A1:P{sheet.max_row}"
    sheet.freeze_panes = "A2"

    summary = workbook.create_sheet("Resumo")
    _append(summary, ["Cajuru A1 — Relatório de auditoria"])
    _append(summary, ["Gerado em", datetime.now().astimezone().strftime("%d/%m/%Y %H:%M")])
    _append(summary, ["READ_ONLY_MODE", "ATIVO"])
    _append(summary, ["Integridade Dropbox", result.safety_message or ("OK" if result.safety_ok else "FALHOU")])
    _append(summary, [])
    _append(summary, ["Métrica", "Valor"])
    for key, value in result.stats.items():
        _append(summary, [key, value])
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 70

    certificates = workbook.create_sheet("Certificados lidos")
    _append(certificates, ["Arquivo", "CNPJ interno", "CNPJ nome", "Aberto", "Chave privada", "Validade", "Ainda não válido", "SHA-256", "Origem candidata", "Código erro", "Erro", "Dropbox"])
    for cert in result.certificados:
        _append(certificates, [
            cert.filename, format_cnpj(cert.cnpj_cert) if cert.cnpj_cert else "",
            format_cnpj(cert.cnpj_filename) if cert.cnpj_filename else "",
            "sim" if cert.opened else "não", "sim" if cert.has_private_key else "não",
            cert.not_after.strftime("%d/%m/%Y") if cert.not_after else "",
            "sim" if cert.not_yet_valid else "não", cert.sha256, cert.password_source or "",
            cert.error_code, cert.error or "", cert.source_path,
        ])

    pdfs = workbook.create_sheet("PDFs de apoio")
    _append(pdfs, ["Arquivo", "Aberto", "Protegido", "Duplicado", "Páginas", "Páginas lidas", "Texto", "OCR", "Documentos encontrados", "Empresas (nome exato)", "Revisão", "Código erro", "SHA-256"])
    for document in result.documents:
        _append(pdfs, [
            document.filename, "sim" if document.opened else "não", "sim" if document.protected else "não",
            "sim" if document.duplicate_sha256 else "não", document.page_count, document.pages_read, "sim" if document.has_text else "não",
            "sim" if document.ocr_used else "não", ", ".join(format_cnpj(item) for item in document.documents),
            ", ".join(document.company_names), document.review_reason, document.error_code, document.sha256,
        ])

    excel_audit = workbook.create_sheet("Auditoria Excel")
    _append(excel_audit, ["GRAVIDADE", "CÓDIGO", "MENSAGEM", "ORIGEM", "ABA", "LINHA/QUANTIDADE"])
    for finding in result.excel_findings:
        _append(excel_audit, [finding.get("severity"), finding.get("code"), finding.get("message"), finding.get("origem"), finding.get("aba"), finding.get("linha", finding.get("quantidade", ""))])

    fd, temporary = tempfile.mkstemp(prefix=f".{destination.name}.", suffix=".xlsx", dir=str(destination.parent))
    os.close(fd)
    try:
        workbook.save(temporary)
        workbook.close()
        os.replace(temporary, destination)
        try:
            destination.chmod(stat.S_IRUSR | stat.S_IWUSR)
        except OSError:
            pass
    finally:
        Path(temporary).unlink(missing_ok=True)
    return destination


def write_html_report(result: PipelineResult, dest: Path) -> Path:
    destination = _guard_destination(result, dest)
    rows: list[str] = []
    for match in result.matches:
        color = "#" + COLORS.get(match.status, "7C8494")
        company = match.cliente.razao_social if match.cliente else "—"
        cnpj = format_cnpj(match.cliente.cnpj) if match.cliente and match.cliente.cnpj else "—"
        filename = match.cert.filename if match.cert else "—"
        rows.append(
            f"<tr><td><span class='badge' style='background:{color}22;color:{color};"
            f"border:1px solid {color}55'>{_esc(match.status)}</span></td>"
            f"<td>{_esc(company)}</td><td class='mono'>{_esc(cnpj)}</td><td class='mono'>{_esc(filename)}</td>"
            f"<td class='muted'>{_esc(match.metodo or '—')}</td><td class='mono'>{match.confianca:.1f}</td>"
            f"<td>{_esc(match.motivo)}</td><td class='muted'>{_esc(' | '.join(match.evidencias))}</td></tr>"
        )
    integrity = result.safety_message or ("Integridade confirmada" if result.safety_ok else "INTEGRIDADE NÃO CONFIRMADA")
    html_document = f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Cajuru A1 — Auditoria</title>
<style>
:root{{--bg:#0A0C11;--surface:#12151C;--surface-2:#171B24;--border:#242A36;--text:#EBEDF1;
--text-muted:#8890A0;--text-faint:#5B6273;--accent:#4C6FE0;}}
*{{box-sizing:border-box}}
body{{font-family:'Segoe UI',-apple-system,Arial,sans-serif;background:var(--bg);color:var(--text);margin:0;font-size:14px;line-height:1.5}}
.mono{{font-family:'JetBrains Mono','Cascadia Code',Consolas,monospace;font-size:12px}}
.muted{{color:var(--text-muted)}}
header{{background:linear-gradient(180deg,var(--surface-2),var(--bg));border-bottom:1px solid var(--border);padding:28px 40px}}
h1{{margin:0 0 6px;font-size:20px;font-weight:650;letter-spacing:-.01em}}
header div{{color:var(--text-muted);font-size:13px}}
.wrap{{padding:28px 40px 60px}}
.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;margin-bottom:22px}}
.card{{background:var(--surface);border:1px solid var(--border);border-radius:12px;padding:16px 18px;position:relative}}
.card::before{{content:"";position:absolute;left:0;top:14px;bottom:14px;width:3px;border-radius:3px;background:var(--accent)}}
.card b{{display:block;font-size:26px;font-weight:650;color:var(--text);font-variant-numeric:tabular-nums}}
.card span{{color:var(--text-muted);font-size:11.5px;text-transform:uppercase;letter-spacing:.4px}}
table{{width:100%;border-collapse:collapse;background:var(--surface);border:1px solid var(--border);border-radius:12px;overflow:hidden}}
th{{background:var(--surface-2);color:var(--text-faint);text-align:left;padding:11px 14px;font-size:10.5px;text-transform:uppercase;letter-spacing:.5px;font-weight:600;border-bottom:1px solid var(--border)}}
td{{padding:11px 14px;border-bottom:1px solid var(--border);font-size:13px;vertical-align:top}}
tr:last-child td{{border-bottom:none}}
tr:hover td{{background:var(--surface-2)}}
.badge{{padding:3px 10px;border-radius:999px;font-size:10.5px;font-weight:700;text-transform:uppercase;letter-spacing:.3px;white-space:nowrap}}
.note{{margin-top:18px;color:var(--text-faint);font-size:12.5px}}
</style></head><body><header><h1>Cajuru A1 — Auditoria e conciliação</h1><div>{_esc(integrity)}</div></header><div class="wrap">
<div class="cards">
<div class="card"><b>{result.stats.get('pfx', 0)}</b><span>PFX/P12</span></div>
<div class="card"><b>{result.stats.get('pdf', 0)}</b><span>PDFs de apoio</span></div>
<div class="card"><b>{result.stats.get('pronto', 0)}</b><span>Prontos (CNPJ interno)</span></div>
<div class="card"><b>{result.stats.get('revisao_manual', 0)}</b><span>Revisão manual</span></div>
<div class="card"><b>{result.stats.get('alterados', 0)}</b><span>Alterações Dropbox</span></div>
</div>
<table><thead><tr><th>Status</th><th>Empresa</th><th>CNPJ</th><th>Arquivo</th><th>Método</th><th>Confiança</th><th>Motivo</th><th>Evidências</th></tr></thead><tbody>{''.join(rows)}</tbody></table>
<p class="note">Nenhum valor de senha é incluído neste relatório. PDFs são evidência de apoio e nunca são enviados como certificado A1.</p></div></body></html>"""
    fd, temporary = tempfile.mkstemp(prefix=f".{destination.name}.", dir=str(destination.parent))
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as stream:
            stream.write(html_document)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, destination)
        try:
            destination.chmod(stat.S_IRUSR | stat.S_IWUSR)
        except OSError:
            pass
    finally:
        Path(temporary).unlink(missing_ok=True)
    return destination


def _esc(value: str) -> str:
    return html.escape(str(value), quote=True)
