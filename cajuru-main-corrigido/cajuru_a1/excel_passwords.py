"""Leitura resiliente e auditável das planilhas de senha."""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from cajuru_a1.cnpjutil import (
    extract_docs_from_text,
    is_valid_doc,
    only_digits,
    pad_cnpj,
)
from cajuru_a1.names import strip_accents
from cajuru_a1.passwords import PasswordEntry, PasswordVault

log = logging.getLogger("cajuru_a1.excel")


def _header_key(value) -> str:
    text = strip_accents("" if value is None else str(value).strip()).upper().replace("_", " ")
    return " ".join("".join(c if c.isalnum() else " " for c in text).split())


NAME_HEADERS = {_header_key(x) for x in ("EMPRESAS", "EMPRESA", "RAZAO", "RAZÃO SOCIAL", "NOME", "CLIENTE", "CLIENTES", "RAZAOSOCIAL", "NOME EMPRESARIAL")}
PASS_HEADERS = {_header_key(x) for x in ("SENHA", "SENHAS", "PASSWORD", "PASSWD", "SENHA CERTIFICADO", "SENHA A1", "SENHA DO CERTIFICADO")}
CNPJ_HEADERS = {_header_key(x) for x in ("CNPJ", "CNPJ/CPF", "CPF", "DOCUMENTO", "CNPJ CPF", "CPF/CNPJ")}
VALID_HEADERS = {_header_key(x) for x in ("VALIDADE", "DATA VALIDADE", "DATA DE VALIDADE", "VENCIMENTO")}


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "SIM" if value else "NÃO"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).strip()


def _columns(row: tuple) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for column, value in enumerate(row):
        header = _header_key(value)
        if header in NAME_HEADERS and "nome" not in mapping:
            mapping["nome"] = column
        elif header in PASS_HEADERS and "senha" not in mapping:
            mapping["senha"] = column
        elif header in CNPJ_HEADERS and "cnpj" not in mapping:
            mapping["cnpj"] = column
        elif header in VALID_HEADERS and "validade" not in mapping:
            mapping["validade"] = column
    return mapping if {"nome", "senha"}.issubset(mapping) else None


def _value(row: tuple, column: int | None) -> object:
    return row[column] if column is not None and column < len(row) else None


def _read_sheet(path: Path, sheet, vault: PasswordVault, say) -> int:
    before = len(vault.entries)
    columns: dict[str, int] | None = None
    header_row = 0
    empty_names = 0
    empty_passwords = 0
    numeric_passwords = 0
    rows_seen = 0

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
        rows_seen = row_number
        if columns is None:
            if row_number <= 100:
                columns = _columns(row)
                if columns:
                    header_row = row_number
                continue
            vault.finding("warning", "cabecalho_ausente", "Aba ignorada: cabeçalho de empresa e senha não encontrado nas primeiras 100 linhas", origem=str(path), aba=sheet.title)
            return 0
        if row_number <= header_row:
            continue
        name_raw = _value(row, columns.get("nome"))
        password_raw = _value(row, columns.get("senha"))
        name = _cell_text(name_raw)
        password = _cell_text(password_raw)
        if not name and not password:
            continue
        if not name:
            empty_names += 1
            continue
        if not password:
            empty_passwords += 1
            continue
        if isinstance(password_raw, (int, float)) and not isinstance(password_raw, bool):
            numeric_passwords += 1

        doc = ""
        doc_raw = _cell_text(_value(row, columns.get("cnpj")))
        if doc_raw:
            digits = only_digits(doc_raw)
            candidate = pad_cnpj(digits)
            if is_valid_doc(candidate):
                doc = candidate
            else:
                docs = extract_docs_from_text(doc_raw)
                if len(docs) == 1:
                    doc = docs[0]
                else:
                    vault.finding("warning", "documento_invalido_excel", "Documento inválido ou ambíguo ignorado", origem=str(path), aba=sheet.title, linha=row_number)
        validity = _cell_text(_value(row, columns.get("validade")))
        vault.add(PasswordEntry(name, password, doc, validity, str(path), sheet.title, row_number))

    if columns is None:
        code = "aba_vazia" if rows_seen == 0 else "cabecalho_ausente"
        message = "Aba vazia ignorada" if rows_seen == 0 else "Aba ignorada: cabeçalho de empresa e senha não encontrado"
        vault.finding("warning", code, message, origem=str(path), aba=sheet.title)
        return 0
    if empty_names:
        vault.finding("info", "nomes_vazios", "Linhas com nome vazio ignoradas", origem=str(path), aba=sheet.title, quantidade=empty_names)
    if empty_passwords:
        vault.finding("warning", "senhas_vazias", "Linhas com senha vazia ignoradas", origem=str(path), aba=sheet.title, quantidade=empty_passwords)
    if numeric_passwords:
        vault.finding("warning", "senhas_numericas", "Senhas armazenadas como número podem ter perdido zeros à esquerda; formate a coluna como texto", origem=str(path), aba=sheet.title, quantidade=numeric_passwords)
    count = len(vault.entries) - before
    say(f"Planilha {path.name} / aba {sheet.title}: {count} linha(s) válida(s)")
    return count


def _read_one(path: Path, vault: PasswordVault, say) -> int:
    before = len(vault.entries)
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        vault.finding("high", "excel_ilegivel", f"Planilha não pôde ser lida ({type(exc).__name__})", origem=str(path))
        say(f"Planilha {path.name}: ERRO de leitura ({type(exc).__name__})")
        return 0
    try:
        for sheet in workbook.worksheets:
            _read_sheet(path, sheet, vault, say)
    finally:
        workbook.close()
    return len(vault.entries) - before


def load_excel_files(paths: list[str | Path], log_fn=None) -> PasswordVault:
    vault = PasswordVault()
    say = log_fn or (lambda message: log.info(message))
    files: list[Path] = []
    for raw in paths or []:
        path = Path(raw).expanduser()
        if path.is_dir():
            files.extend(sorted(item for item in path.iterdir() if item.suffix.lower() in {".xlsx", ".xlsm"}))
        else:
            files.append(path)

    seen: set[str] = set()
    for path in files:
        key = str(path.resolve(strict=False)).casefold()
        if key in seen:
            vault.finding("info", "arquivo_duplicado", "A mesma planilha foi configurada mais de uma vez", origem=str(path))
            continue
        seen.add(key)
        if not path.exists() or not path.is_file():
            vault.finding("high", "excel_ausente", "Planilha não encontrada", origem=str(path))
            say(f"Planilha NÃO encontrada: {path}")
            continue
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            vault.finding("warning", "formato_nao_suportado", "Arquivo ignorado: formato não suportado", origem=str(path))
            continue
        count = _read_one(path, vault, say)
        say(f"Planilha {path.name}: +{count} registro(s)")
    say(f"Cofre em memória: {len(vault.entries)} registro(s), {len(vault.all_unique)} senha(s) única(s).")
    return vault
